from openpyxl import Workbook, load_workbook # type: ignore
from openpyxl.styles import Font, Alignment # type: ignore
from pathlib import Path

FILE = "nhanvien.xlsx"
SHEET = "NhanVien"

def init_file():
    """Tạo file + header nếu chưa tồn tại."""
    if Path(FILE).exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET


    headers = ["STT", "Mã", "Tên", "Tuổi"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 6   
    ws.column_dimensions['B'].width = 12  
    ws.column_dimensions['C'].width = 22  
    ws.column_dimensions['D'].width = 10  

    wb.save(FILE)

def open_ws():
    wb = load_workbook(FILE)
    return wb, wb[SHEET]

def add_employee(ma: str, ten: str, tuoi: int):
    wb, ws = open_ws()
    
    stt = ws.max_row
    ws.append(["", ma.strip(), ten.strip(), int(tuoi)])
    
    renumber(ws)
    wb.save(FILE)

def renumber(ws):
    """Đánh lại STT theo thứ tự từ 1..n ở cột A."""
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        row[0].value = i

def list_all():
    wb, ws = open_ws()
    print(f"{'STT':<5}{'Mã':<10}{'Tên':<20}{'Tuổi':<5}")
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] is None:
            continue
        stt, ma, ten, tuoi = r
        print(f"{(stt or ''):<5}{ma:<10}{ten:<20}{tuoi:<5}")
    wb.close()

def sort_by_age_asc():
    wb, ws = open_ws()

    data = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] is None:
            continue
        _, ma, ten, tuoi = r
        data.append([ma, ten, int(tuoi)])


    data.sort(key=lambda x: x[2])


    ws.delete_rows(2, ws.max_row)

    for i, (ma, ten, tuoi) in enumerate(data, start=1):
        ws.append([i, ma, ten, tuoi])

    wb.save(FILE)
    wb.close()
    print("Đã sắp xếp theo tuổi tăng dần.")

def main():
    init_file()
    while True:
        print("\n===== QUẢN LÝ NHÂN VIÊN (Excel) =====")
        print("1. Thêm nhân viên")
        print("2. Xem danh sách")
        print("3. Sắp xếp theo Tuổi tăng dần")
        print("0. Thoát")
        chon = input("Chọn: ").strip()

        if chon == "1":
            ma = input("Mã: ")
            ten = input("Tên: ")
            while True:
                try:
                    tuoi = int(input("Tuổi: "))
                    break
                except ValueError:
                    print("Tuổi phải là số nguyên, nhập lại nhé!")
            add_employee(ma, ten, tuoi)
            print("✔ Đã lưu vào Excel.")
        elif chon == "2":
            list_all()
        elif chon == "3":
            sort_by_age_asc()
        elif chon == "0":
            print("Tạm biệt!")
            break
        else:
            print("Không hợp lệ, chọn lại nhé!")

if __name__ == "__main__":
    main()